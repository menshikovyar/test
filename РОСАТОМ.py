import requests
import pandas as pd
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Функция для получения HTML-контента страницы
def get_html_content(url):
    response = requests.get(url)
    return response.content

# Функция для извлечения данных о курсах валют из HTML-контента
def extract_currency_data(html_content):
    # Создаем объект BeautifulSoup для парсинга HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # Находим элементы, содержащие данные о курсах USD-RUB
    usd_data_elements = soup.find_all('div', class_='usd-data')

    # Находим элементы, содержащие данные о курсах JPY-RUB
    jpy_data_elements = soup.find_all('div', class_='jpy-data')

    # Создаем пустые списки для хранения данных
    usd_dates = []
    usd_rates = []
    usd_times = []

    jpy_dates = []
    jpy_rates = []
    jpy_times = []

    # Извлекаем данные о курсах USD-RUB
    for element in usd_data_elements:
        usd_dates.append(element.find('span', class_='date').text.strip())
        usd_rates.append(element.find('span', class_='rate').text.strip())
        usd_times.append(element.find('span', class_='time').text.strip())

    # Извлекаем данные о курсах JPY-RUB
    for element in jpy_data_elements:
        jpy_dates.append(element.find('span', class_='date').text.strip())
        jpy_rates.append(element.find('span', class_='rate').text.strip())
        jpy_times.append(element.find('span', class_='time').text.strip())

    return usd_dates, usd_rates, usd_times, jpy_dates, jpy_rates, jpy_times


# Функция для получения данных о курсах валют по API MOEX ISS
def get_currency_rates(currency, start_date, end_date):
    url = f"https://iss.moex.com/iss/history/engines/currency/markets/selt/boards/CETS/securities/{currency}.json"
    params = {
        "iss.only": "history",
        "history.columns": "TRADEDATE, CLOSE, TRADETIME",
        "from": start_date,
        "till": end_date
    }
    response = requests.get(url, params=params)
    data = response.json()
    return data["history"]["data"]


# Функция для добавления данных в существующий файл Excel
def append_to_excel(dataframe, file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    start_row = ws.max_row + 1
    for row in dataframe.itertuples(index=False):
        ws.append(row)
    wb.save(file_name)


# Получаем HTML-контент страницы с курсами валют
url = "https://www.moex.com/ru/derivatives/currency-rate.aspx?currency=USD_RUB"
html_content = get_html_content(url)

# Извлекаем данные о курсах валют из HTML-контента
usd_dates, usd_rates, usd_times, jpy_dates, jpy_rates, jpy_times = extract_currency_data(html_content)

# Получаем текущую дату и дату начала предыдущего месяца
end_date = datetime.now().date()
start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)

# Получаем данные о курсах USD/RUB за предыдущий месяц
usd_data = get_currency_rates("USD_RUB", start_date=start_date, end_date=end_date)

# Получаем данные о курсах JPY/RUB за предыдущий месяц
jpy_data = get_currency_rates("JPY_RUB", start_date=start_date, end_date=end_date)

# Создаем пустой DataFrame для хранения всех данных
merged_df = pd.DataFrame(columns=["Дата USD/RUB", "Курс USD/RUB", "Время USD/RUB",
                                  "Дата JPY/RUB", "Курс JPY/RUB", "Время JPY/RUB"])

# Заполняем DataFrame данными для USD/RUB из HTML-контента
for i in range(len(usd_dates)):
    merged_df.loc[i, "Дата USD/RUB"] = usd_dates[i]
    merged_df.loc[i, "Курс USD/RUB"] = usd_rates[i]
    merged_df.loc[i, "Время USD/RUB"] = usd_times[i]

# Заполняем DataFrame данными для JPY/RUB из HTML-контента
for i in range(len(jpy_dates)):
    merged_df.loc[i, "Дата JPY/RUB"] = jpy_dates[i]
    merged_df.loc[i, "Курс JPY/RUB"] = jpy_rates[i]
    merged_df.loc[i, "Время JPY/RUB"] = jpy_times[i]

# Добавляем столбец для результата деления курса USD/RUB на JPY/RUB
merged_df["Результат"] = merged_df["Курс USD/RUB"].astype(float) / merged_df["Курс JPY/RUB"].astype(float)

# Добавляем новые данные в файл Excel без удаления существующих данных
file_name = 'currency_rates.xlsx'
append_to_excel(merged_df, file_name)

